import openpyxl 
from ..fitting.get_objective_function import get_objective_function
from ..misc.write_spreadsheet import write_spreadsheet

def export_label_results(label_model,fn="output.xlsx",show_chi=True,show_emu=True,show_fluxes=True):
  a,chi_dict_not_rsm,simulation_not_rsm=get_objective_function(label_model,output=False,rsm="never")
  a,chi_dict_with_rsm,simulation_with_rsm=get_objective_function(label_model,output=False,rsm="always")
  sheet_row_data_dict={}
  order_vector=[]
  for n_condition,condition in enumerate(label_model.initial_label):
      order_vector.append(condition)
      sheet_row_data_dict[condition]=[["Metabolite/s","Isotopologue","Simulated value","Experimental Mean","Experimental SD"]]
      if show_chi==True:
         sheet_row_data_dict[condition][0].append("Chi Square") 
      row_n=1
      for emu in label_model.experimental_dict[condition]:
         rsm_show_chi=True
         for mi in sorted(label_model.experimental_dict[condition][emu]):
                if mi==0:
                   if simulation_not_rsm[condition][emu][mi]>label_model.experimental_dict[condition][emu][mi]["m"]:
                      rsm_show_chi=False
                metabolite_name=emu[4:]
                m_name="m"+str(mi)
                """sim=label_model.condition_simulation_results_dict[condition][emu][mi]
                      try:   
                        exp_m=label_model.experimental_dict[condition][emu][mi]["m"]/(1-label_model.experimental_dict[condition][emu][0]["m"])
                        exp_sd=max(label_model.experimental_dict[condition][emu][mi]["sd"]/(1-label_model.experimental_dict[condition][emu][0]["m"]),label_model.minimum_sd) 
                      except:
                        exp_m=0
                        exp_sd=0 
                      chi=round(pow((sim-exp_m)/exp_sd,2),3)
                else:"""
                sim=simulation_not_rsm[condition][emu][mi]   
                exp_m=label_model.experimental_dict[condition][emu][mi]["m"]
                exp_sd=label_model.experimental_dict[condition][emu][mi]["sd"] 
                chi=chi_dict_not_rsm[condition][emu][mi]   #round(pow((sim-exp_m)/exp_sd,2),3)
                if emu in label_model.data_name_emu_dict:
                   metabolite_name=label_model.data_name_emu_dict[emu]
                row_n+=1
                row=[metabolite_name,m_name,round(sim,4),round(exp_m,4),round(exp_sd,4)]
                if show_chi and not (emu in label_model.rsm_list and rsm_show_chi):
                   row.append(round(chi,4))
                sheet_row_data_dict[condition].append(row)
         sheet_row_data_dict[condition].append([])
         if emu in label_model.rsm_list:
            
            for mi in sorted(label_model.experimental_dict[condition][emu]):
                if mi==0:
                   continue 
                metabolite_name=emu[4:]
                m_name="m"+str(mi)+"/Sm"
                sim=simulation_with_rsm[condition][emu][mi]   
                exp_m=label_model.experimental_dict[condition][emu][mi]["m"]/(1-label_model.experimental_dict[condition][emu][0]["m"])
                exp_sd=label_model.experimental_dict[condition][emu][mi]["sd"]/(1-label_model.experimental_dict[condition][emu][0]["m"]) 
                
                if emu in label_model.data_name_emu_dict:
                   metabolite_name=label_model.data_name_emu_dict[emu]
                row_n+=1
                row=[metabolite_name,m_name,round(sim,4),round(exp_m,4),round(exp_sd,4)]
                if show_chi and rsm_show_chi:
                   chi=chi_dict_with_rsm[condition][emu][mi]   #round(pow((sim-exp_m)/exp_sd,2),3)
                   row.append(round(chi,4))
                sheet_row_data_dict[condition].append(row)
         sheet_row_data_dict[condition].append([])             
      if show_emu:                 
         for size in label_model.size_emu_c_eqn_dict:
           sheet=condition+" emu size%s"%(size)
           order_vector.append(sheet)
           sheet_row_data_dict[sheet]=[]
           for n,x in enumerate(label_model.condition_size_yy_dict[condition][size]):
                sheet_row_data_dict[sheet].append([label_model.size_inverse_variable_dict[size][n],x])
  if show_fluxes:
     row=["ID","Name","Stoichiometry","Value"]
     sheet_row_data_dict["Reaction fluxes"]=[row]
     order_vector.append("Reaction fluxes")   
     for x in sorted(label_model.constrained_model.solution.x_dict,key=lambda v: v.upper()):  
         if "_RATIO" in x:
            continue
         reaction=label_model.constrained_model.reactions.get_by_id(x)
         sheet_row_data_dict["Reaction fluxes"].append([reaction.id,reaction.name,reaction.reaction,label_model.constrained_model.solution.x_dict[x]])
  write_spreadsheet(file_name=fn,sheet_row_data_dict=sheet_row_data_dict,sheet_order=order_vector)    
  a,b,c=get_objective_function(label_model,force_balance=label_model.force_balance,output=False,rsm="dynamic")
  #wb = openpyxl.Workbook()
  """for n_condition,condition in enumerate(label_model.initial_label):
    if n_condition==0:
       sheet = wb.active
    else:
       sheet=wb.create_sheet(title=condition)
    sheet.title = condition
    sheet['A1']="Metabolite/s"
    sheet['B1']="Isotopologue"
    sheet['C1']="Simulated value"
    sheet['D1']="Experimental Mean"
    sheet['E1']="Experimental SD"
    if show_chi==True:
       sheet['F1']="Chi Square"
    row_n=1
    for emu in label_model.experimental_dict[condition]:
         for mi in label_model.experimental_dict[condition][emu]:
                
                metabolite_name=emu[4:]
                m_name="m"+str(mi)
                if emu in label_model.rsm_list:
                      if mi==0:
                         continue
                      m_name+="/mS"
                      sim=label_model.condition_simulation_results_dict[condition][emu][mi]
                      try:   
                        exp_m=label_model.experimental_dict[condition][emu][mi]["m"]/(1-label_model.experimental_dict[condition][emu][0]["m"])
                        exp_sd=label_model.experimental_dict[condition][emu][mi]["sd"]/(1-label_model.experimental_dict[condition][emu][0]["m"]) 
                      except:
                        exp_m=0
                        exp_sd=0 
                      chi=round(pow((sim-exp_m)/exp_sd,2),3)
                else:
                  sim=label_model.condition_simulation_results_dict[condition][emu][mi]   
                  exp_m=label_model.experimental_dict[condition][emu][mi]["m"]
                  exp_sd=label_model.experimental_dict[condition][emu][mi]["sd"] 
                  chi=round(pow((sim-exp_m)/exp_sd,2),3)
                if emu in label_model.data_name_emu_dict:
                   metabolite_name=label_model.data_name_emu_dict[emu]
                row_n+=1
                label_model.data_name_emu_dict
                sheet['A'+str(row_n)]=metabolite_name
                sheet['B'+str(row_n)]=m_name
                sheet['C'+str(row_n)]=str(round(sim,4))
                sheet['D'+str(row_n)]=str(round(exp_m,4))
                sheet['E'+str(row_n)]=str(round(exp_sd,4))
                if show_chi==True:
                   sheet['F'+str(row_n)]=str(round(chi,4))
         row_n+=1
    for size in label_model.size_emu_c_eqn_dict:
           sheet=wb.create_sheet(title=condition+" emu size%s"%(size))
           for n,x in enumerate(label_model.condition_size_yy_dict[condition][size]):
                sheet['A'+str(n+1)]=label_model.size_inverse_variable_dict[size][n]
                sheet['B'+str(n+1)]=str(round(x,8)) 
  wb.save(fn)"""

